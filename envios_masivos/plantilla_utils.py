import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def crear_plantilla_excel(nombre_archivo='DEVOLVER_CORREOS.xlsx'):
    columnas = [
        'FICHA',
        'NIVEL DE FORMACION',
        'NOMBRES Y APELLIDOS',
        'CORREO',
        'CERTIFICADO ETAPA PRODUCTIVA',
        'EVALUACION PARCIAL',
        'EVALUACION FINAL',
        'TYT (TECNOLOGOS)',
        'CERTIFICADO VIGENCIA',
        'CERTIFICADO AGENCIA DE EMPLEO SENA',
        'CARNET DESTRUIDO',
        'PAZ Y SALVO ACADEMICO ADMINISTRATIVO'
    ]

    # Crear y guardar la plantilla con pandas
    df = pd.DataFrame(columns=columnas)
    df.to_excel(nombre_archivo, index=False)

    # Abrir con openpyxl para modificar el ancho de columnas
    wb = load_workbook(nombre_archivo)
    ws = wb.active

    # Ajustar ancho de cada columna basado en el largo del título
    for i, col in enumerate(columnas, 1):
        max_length = len(col) + 2  # margen extra
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = max_length

    wb.save(nombre_archivo)
    print(f"✅ Plantilla Excel creada correctamente: {nombre_archivo}")
